# -*- coding: utf-8 -*-
import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.cell import get_column_letter

def select_old_file():
    path = filedialog.askopenfilename(title="选择历史版清单", filetypes=[("Excel 文件", "*.xlsx;*.xls")])
    if path:
        entry_old.delete(0, tk.END)
        entry_old.insert(0, path)

def select_new_file():
    path = filedialog.askopenfilename(title="选择最新版清单", filetypes=[("Excel 文件", "*.xlsx;*.xls")])
    if path:
        entry_new.delete(0, tk.END)
        entry_new.insert(0, path)

def select_save_path():
    path = filedialog.asksaveasfilename(
        title="选择保存位置", 
        defaultextension=".xlsx", 
        filetypes=[("Excel 文件", "*.xlsx")],
        initialfile="数据核对_红底高亮结果.xlsx"
    )
    if path:
        entry_save.delete(0, tk.END)
        entry_save.insert(0, path)

def start_comparison():
    old_file = entry_old.get().strip()
    new_file = entry_new.get().strip()
    output_file = entry_save.get().strip()

    if not old_file or not new_file or not output_file:
        messagebox.showwarning("提示窗口", "请先把【旧版】、【新版】和【保存路径】都选择完毕！")
        return

    # 按钮防手抖，提示正在处理
    btn_start.config(text="正在拼命核对中...", state=tk.DISABLED)
    root.update()

    try:
        wb_old = load_workbook(old_file, data_only=True)
        wb_new = load_workbook(new_file)
    except Exception as e:
        messagebox.showerror("读取失败", f"文件可能被占用，请关闭相关Excel后重试！\n\n报错: {e}")
        reset_button()
        return

    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    diff_records = []

    for sheet_name in wb_new.sheetnames:
        if sheet_name not in wb_old.sheetnames:
            continue

        ws_old = wb_old[sheet_name]
        ws_new = wb_new[sheet_name]

        max_row = max(ws_old.max_row, ws_new.max_row)
        max_col = max(ws_old.max_column, ws_new.max_column)

        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                val_old = ws_old.cell(row=row, column=col).value
                val_new = ws_new.cell(row=row, column=col).value

                str_old = str(val_old).strip() if val_old is not None else ""
                str_new = str(val_new).strip() if val_new is not None else ""

                if str_old != str_new:
                    ws_new.cell(row=row, column=col).fill = red_fill
                    diff_records.append({
                        "工作表名": sheet_name,
                        "单元格": f"{get_column_letter(col)}{row}",
                        "历史版数据": val_old,
                        "最新版数据": val_new
                    })

    if diff_records:
        df_diff = pd.DataFrame(diff_records)
        ws_report = wb_new.create_sheet(title="数据差异明细汇总")
        for c_idx, col_name in enumerate(df_diff.columns, 1):
            ws_report.cell(row=1, column=c_idx, value=col_name)
        for r_idx, row_data in enumerate(df_diff.values, 2):
            for c_idx, value in enumerate(row_data, 1):
                ws_report.cell(row=r_idx, column=c_idx, value=value)
        msg = f"核对完毕！共标红 {len(diff_records)} 处差异。\n明细表已生成在最后一个Sheet页。"
    else:
        msg = "核对完毕！两份清单数据完全一致，无任何变动。"

    try:
        wb_new.save(output_file)
        messagebox.showinfo("核对成功", msg)
    except Exception as e:
        messagebox.showerror("保存失败", f"保存失败，请检查结果文件是否正被打开！\n\n报错: {e}")
    finally:
        reset_button()

def reset_button():
    btn_start.config(text="开始执行比对", state=tk.NORMAL)

# ================= 构建可视化操作面板 =================
root = tk.Tk()
root.title("Excel 两表差异自动高亮核对助手")
root.geometry("550x260")
root.resizable(False, False)

# 统一边距配置
padx, pady = 15, 10

# 1. 历史版本选择行
tk.Label(root, text="历史版清单 (旧):").grid(row=0, column=0, padx=padx, pady=pady, sticky="e")
entry_old = tk.Entry(root, width=45)
entry_old.grid(row=0, column=1, pady=pady)
tk.Button(root, text="浏览...", command=select_old_file).grid(row=0, column=2, padx=padx, pady=pady)

# 2. 最新版本选择行
tk.Label(root, text="最新版清单 (新):").grid(row=1, column=0, padx=padx, pady=pady, sticky="e")
entry_new = tk.Entry(root, width=45)
entry_new.grid(row=1, column=1, pady=pady)
tk.Button(root, text="浏览...", command=select_new_file).grid(row=1, column=2, padx=padx, pady=pady)

# 3. 保存路径选择行
tk.Label(root, text="比对结果保存至:").grid(row=2, column=0, padx=padx, pady=pady, sticky="e")
entry_save = tk.Entry(root, width=45)
entry_save.grid(row=2, column=1, pady=pady)
tk.Button(root, text="选择...", command=select_save_path).grid(row=2, column=2, padx=padx, pady=pady)

# 4. 执行按钮
btn_start = tk.Button(root, text="开始执行比对", bg="#4CAF50", fg="black", font=("Arial", 12, "bold"), width=20, command=start_comparison)
btn_start.grid(row=3, column=0, columnspan=3, pady=20)

if __name__ == "__main__":
    root.mainloop()